
import os
import pandas as pd
from openpyxl import load_workbook, Workbook

def write_into_excel(nodes, llama2_responses, langchain_responses, questions, file_path, ):#llama2_code_responses = []):
    
    # Create a Workbook if the file is not present
    if(os.path.isfile(file_path)):
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()

    sheet = workbook.active

    # To append into the exisiting Workbook if it has some data already
    sheet.cell(row=1, column=1, value="Question")
    sheet.cell(row=1, column=2, value="Llama 2 - Together ai")
    sheet.cell(row=1, column=3, value="Langchain")
    next_row = sheet.max_row + 2

    for index, (question, ll2_response, lg_response) in enumerate(zip(questions,llama2_responses,langchain_responses),start=next_row):
        sheet.cell(row=index, column=1, value=question)
        sheet.cell(row=index, column=2, value=ll2_response)
        sheet.cell(row=index, column=3, value=lg_response)
       # sheet.cell(row=index, column=4, value=llama2_code_response)

    # Save the workbook
    workbook.save(file_path)

def write_into_excel_3Models(nodes, llama2_responses, langchain_responses, questions, file_path, llama2_code_responses = []):
    
    # Create a Workbook if the file is not present
    if(os.path.isfile(file_path)):
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()

    sheet = workbook.active

    # To append into the exisiting Workbook if it has some data already
    sheet.cell(row=1, column=1, value="Question")
    sheet.cell(row=1, column=2, value="Llama 2 - Together ai")
    sheet.cell(row=1, column=3, value="Langchain")
    next_row = sheet.max_row + 2

    for index, (question, ll2_response, lg_response,llama2_code_response) in enumerate(zip(questions,llama2_responses,langchain_responses, llama2_code_responses),start=next_row):
        sheet.cell(row=index, column=1, value=question)
        sheet.cell(row=index, column=2, value=ll2_response)
        sheet.cell(row=index, column=3, value=lg_response)
        sheet.cell(row=index, column=4, value=llama2_code_response)

    # Save the workbook
    workbook.save(file_path)